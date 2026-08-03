import os
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
from LX_BaseFrame import LX_BaseFrame, LXLockState
from LX_CommCentre import LX_CommCentre

class F_MIR(LX_BaseFrame):
    def __init__(self):
        super().__init__(alias="MIR")

    def execute(self, row_data: dict) -> bool:
        if str(row_data.get("LX", "")).upper() == "X":
            self.set_lock_status(LXLockState.SKIP)
            return True

        self.set_lock_status(LXLockState.BUSY)
        log = row_data.get("_log_callback", print)
        log(f"[{self.alias}] External Lock ENGAGED. Mirroring cells (Reading via openpyxl, Writing via xlsxwriter)...")

        try:
            dest_sheet_name = row_data.get("D_Sheet", "Sheet1")
            ws_dest = row_data.get("_ws_dict").get(dest_sheet_name)
            wb_xlsx = row_data.get("_wb")

            source_path = LX_CommCentre.resolve_path(row_data.get("S_Path", ""))
            source_file = row_data.get("S_File_Name", "")
            source_full_path = os.path.join(source_path, source_file)
            source_sheet_name = row_data.get("S_Sheet", "Sheet1")

            start_cell = row_data.get("S_Start_Cell_Range", "")
            end_cell = row_data.get("S_End_Cell_Range", "")
            dest_start_cell = row_data.get("D_Start_Cell_Range", "")

            if not os.path.exists(source_full_path):
                log(f"[{self.alias}] WARNING: Source file for mirror not found: {source_full_path}")
                self.set_lock_status(LXLockState.FREE)
                return True

            wb_source = openpyxl.load_workbook(source_full_path, data_only=True)
            if source_sheet_name not in wb_source.sheetnames:
                raise ValueError(f"Sheet {source_sheet_name} not found in {source_full_path}")
            ws_source = wb_source[source_sheet_name]

            sc_col, sc_row = coordinate_from_string(start_cell)
            ec_col, ec_row = coordinate_from_string(end_cell)
            sc_col_idx = column_index_from_string(sc_col)
            ec_col_idx = column_index_from_string(ec_col)
            dc_col, dc_row = coordinate_from_string(dest_start_cell)
            dc_col_idx = column_index_from_string(dc_col)

            for r_offset in range(ec_row - sc_row + 1):
                for c_offset in range(ec_col_idx - sc_col_idx + 1):
                    src_col_letter = get_column_letter(sc_col_idx + c_offset)
                    src_cell_coord = f"{src_col_letter}{sc_row + r_offset}"
                    
                    dst_col_letter = get_column_letter(dc_col_idx + c_offset)
                    dst_cell_coord = f"{dst_col_letter}{dc_row + r_offset}"
                    
                    src_cell = ws_source[src_cell_coord]
                    
                    # For a perfect implementation, formats should be translated from openpyxl to xlsxwriter.
                    # Since we are writing using xlsxwriter, we write the value. Image boundary protection is implicit
                    # because xlsxwriter handles images in a separate drawing layer.
                    
                    if src_cell.value is not None:
                        # Translate basic formats (optional step depending on complexity desired)
                        fmt_props = {}
                        if src_cell.has_style:
                            if src_cell.font.bold: fmt_props['bold'] = True
                            if src_cell.font.italic: fmt_props['italic'] = True
                            if src_cell.font.color and src_cell.font.color.type == 'rgb':
                                # openpyxl RGB is usually ARGB (e.g. 00FF0000). xlsxwriter needs #RRGGBB
                                rgb = str(src_cell.font.color.rgb)
                                if len(rgb) == 8:
                                    fmt_props['font_color'] = '#' + rgb[2:]
                            
                        cell_format = wb_xlsx.add_format(fmt_props) if fmt_props else None
                        
                        if cell_format:
                            ws_dest.write(dst_cell_coord, src_cell.value, cell_format)
                        else:
                            ws_dest.write(dst_cell_coord, src_cell.value)
                        
            log(f"[{self.alias}] Mirrored {start_cell}:{end_cell} to {dest_start_cell}")
            wb_source.close()

            self.set_lock_status(LXLockState.FREE)
            return True

        except Exception as err:
            log(f"[{self.alias}] CRITICAL ERROR: {err}")
            self.set_lock_status(LXLockState.ERROR)
            return False
